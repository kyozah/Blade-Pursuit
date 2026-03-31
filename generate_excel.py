import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Tạo workbook mới
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Định nghĩa các cột
headers = [
    "Test Case ID",
    "Test Case Title", 
    "Expected Result",
    "Actual Result",
    "Run Type",
    "Tested By",
    "Test Step Detail",
    "Status"
]

# Thêm header
ws.append(headers)

# Định dạng header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Dữ liệu test cases (Movement Tests)
movement_tests = [
    ("1.1", "Player_Moves_Forward_On_W_Input", "Z position increases", 
     "Assert.Greater(playerObject.transform.position.z, initialPosition.z)", 
     "Automatic", "Phất", "Move player forward using W key and verify Z position change", "Passed"),
    
    ("1.2", "Player_Moves_Backward_On_S_Input", "Z position decreases",
     "Assert.Less(playerObject.transform.position.z, initialPosition.z)",
     "Automatic", "Phất", "Move player backward using S key and verify Z position change", "Passed"),
    
    ("1.3", "Player_Moves_Left_And_Right", "X position changes correctly",
     "Assert.Greater/Less(playerObject.transform.position.x, initialPosition.x)",
     "Automatic", "Phất", "Move player left (A) and right (D) and verify X position changes", "Passed"),
    
    ("1.4", "Player_Can_Roll_Dodge_Quickly", "Distance traveled >= 2m in 0.5s",
     "Assert.GreaterOrEqual(distanceTraveled, expectedMinDistance * 0.8f)",
     "Automatic", "Phất", "Execute dodge roll and measure travel distance", "Passed"),
    
    ("1.5", "Movement_Velocity_Is_Valid", "Velocity between 0 and maxSpeed",
     "Assert.Greater(currentVelocity.magnitude, 0f) AND Assert.Less(currentVelocity.magnitude, maxAllowedSpeed)",
     "Automatic", "Phất", "Check velocity is within valid range (0 < v < 20)", "Passed"),
    
    ("1.6", "Player_Doesnt_Move_Without_Input", "Position unchanged with no input",
     "Assert.AreEqual(initialPosition, finalPosition)",
     "Automatic", "Phất", "Verify player stays in place when no movement input", "Passed"),
    
    ("1.7", "Collision_Blocks_Movement", "Player blocked by wall collision",
     "Assert.Less(distanceMoved, 50f)",
     "Automatic", "Phất", "Create wall collider and verify player cannot pass through", "Passed"),
    
    ("1.8", "Sprint_Increases_Speed", "Sprint distance > normal distance",
     "Assert.Greater(sprintDistance, normalDistance)",
     "Automatic", "Phất", "Compare movement distance with and without sprint", "Passed"),
    
    ("1.9", "Player_Can_Jump", "Y position increases when jumping",
     "Assert.Greater(jumpHeight, 0f)",
     "Automatic", "Phất", "Execute jump and verify Y axis position increase", "Passed"),
    
    ("1.10", "Knockback_Effect", "Knockback distance >= 2m",
     "Assert.Greater(knockbackDistance, 2f)",
     "Automatic", "Phất", "Apply knockback force and measure distance traveled", "Passed"),
]

# UI Tests
ui_tests = [
    ("2.1", "HealthBar_Displays_Correctly", "Health bar visible with fillAmount=1.0",
     "Assert.IsTrue(healthBarObject.activeInHierarchy) AND Assert.AreEqual(healthBarImage.fillAmount, 1f)",
     "Automatic", "Phất", "Verify health bar displays correctly at game start", "Passed"),
    
    ("2.2", "Death_Menu_Appears_On_Player_Death", "Death menu appears when HP=0",
     "Assert.IsTrue(deathMenuObject.activeInHierarchy) AND Assert.Greater(canvasGroup.alpha, 0f)",
     "Automatic", "Phất", "Simulate player death and verify menu appears", "Passed"),
    
    ("2.3", "Victory_Menu_Appears_On_Level_Complete", "Victory menu visible when level complete",
     "Assert.IsTrue(victoryMenuObject.activeInHierarchy) AND Assert.Greater(canvasGroup.alpha, 0f)",
     "Automatic", "Phất", "Simulate level completion and verify victory menu shows", "Passed"),
    
    ("2.4", "HealthBar_Updates_On_Health_Change", "fillAmount changes with HP",
     "Assert.AreEqual(healthBarImage.fillAmount, 0.5f)",
     "Automatic", "Phất", "Modify health and verify bar updates in real-time", "Passed"),
    
    ("2.5", "Damage_Popup_Appears", "Damage text popup shows on damage",
     "Assert.IsTrue(damagePopup.activeInHierarchy) AND Assert.AreEqual(damageText.text, \"-20\")",
     "Automatic", "Phất", "Trigger damage and verify damage number appears", "Passed"),
    
    ("2.6", "Score_Display", "Score text updates correctly",
     "Assert.AreEqual(scoreText.text, \"Score: 100\")",
     "Automatic", "Phất", "Update score and verify display text changes", "Passed"),
    
    ("2.7", "Pause_Menu_Toggle", "Pause menu can toggle on/off",
     "Assert.IsTrue/IsFalse(pauseMenu.activeInHierarchy)",
     "Automatic", "Phất", "Toggle pause menu and verify visibility state", "Passed"),
    
    ("2.8", "Button_Interactions", "Button responds to click event",
     "Assert.IsTrue(isClicked)",
     "Automatic", "Phất", "Click button and verify onClick event triggers", "Passed"),
    
    ("2.9", "Text_Font_And_Size", "Text has correct font properties",
     "Assert.AreEqual(text.fontSize, 30) AND Assert.AreEqual(text.fontStyle, FontStyle.Bold)",
     "Automatic", "Phất", "Verify text font size (30) and style (Bold)", "Passed"),
    
    ("2.10", "GameOver_Screen_Delay", "Screen appears after 2s delay",
     "Assert.IsTrue(gameOverScreen.activeInHierarchy) AND Assert.GreaterOrEqual(elapsedTime, deathDelay)",
     "Automatic", "Phất", "Wait for game over delay and verify screen appears", "Passed"),
]

# Health System Tests
health_tests = [
    ("3.1", "Health_Decreases_On_Damage", "HP decreases by damage amount",
     "Assert.AreEqual(healthAfterDamage, initialHealth - damageAmount)",
     "Automatic", "Phất", "Apply 20 damage and verify HP: 100->80", "Passed"),
    
    ("3.2", "Health_Increases_With_Healing_Item", "HP increases but not exceed max",
     "Assert.Greater(healthAfterHealing, healthAfterDamage) AND Assert.LessOrEqual(healthAfterHealing, maxHealth)",
     "Automatic", "Phất", "Apply healing and verify HP increases to max 100", "Passed"),
    
    ("3.3", "Player_Dies_When_Health_Reaches_Zero", "isDead=true when HP=0",
     "Assert.LessOrEqual(healthAfterCriticalDamage, 0f) AND Assert.IsTrue(isDead) AND Assert.IsFalse(isAlive)",
     "Automatic", "Phất", "Deal massive damage and verify death state", "Passed"),
    
    ("3.4", "Invincibility_Frames", "No damage taken during i-frames",
     "Assert.AreEqual(healthBefore, healthAfter)",
     "Automatic", "Phất", "Apply damage during invincibility and verify HP unchanged", "Passed"),
    
    ("3.5", "Knockback_Damage", "Damage applied with knockback force",
     "Assert.Less(healthAfter, initialHealth)",
     "Automatic", "Phất", "Apply knockback damage and verify HP decreases", "Passed"),
    
    ("3.6", "Max_Health_Cap", "HP cannot exceed max limit",
     "Assert.AreEqual(currentHealth, maxHealth)",
     "Automatic", "Phất", "Try to heal 500 and verify HP capped at 100", "Passed"),
    
    ("3.7", "Different_Damage_Sources", "Multiple damage sources calculated correctly",
     "Assert.AreEqual(afterEnvDamage, initialHealth - enemyDamage - environmentDamage)",
     "Automatic", "Phất", "Apply damage from enemy (15) + environment (10) = 25 total", "Passed"),
    
    ("3.8", "Damage_Over_Time", "DoT accumulates damage over time",
     "Assert.Greater(totalDamage, 5f)",
     "Automatic", "Phất", "Apply DoT for 2 seconds and verify ~10 damage accumulated", "Passed"),
    
    ("3.9", "Armor_Damage_Reduction", "Armor reduces damage received",
     "Assert.Less(damageWithArmor, damageWithoutArmor)",
     "Automatic", "Phất", "Compare damage with/without 10% armor mitigation", "Passed"),
    
    ("3.10", "Respawn_System", "Player respawns at spawn point with full HP",
     "Assert.IsTrue(isAlive) AND Assert.AreEqual(currentHealth, maxHealth)",
     "Automatic", "Phất", "Kill player then respawn and verify HP=100", "Passed"),
]

# Thêm dữ liệu vào sheet
test_counter = 1
for test in movement_tests + ui_tests + health_tests:
    row_data = [
        test_counter,  # Test Case ID
        test[1],       # Test Case Title
        test[2],       # Expected Result
        test[3],       # Actual Result / Code
        test[4],       # Run Type
        test[5],       # Tested By
        test[6],       # Test Step Detail
        test[7]        # Status
    ]
    
    ws.append(row_data)
    
    # Định dạng hàng
    for i, cell in enumerate(ws[test_counter + 1]):
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if test[7] == "Passed":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    test_counter += 1

# Điều chỉnh độ rộng cột
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 12

# Đặt chiều cao hàng header
ws.row_dimensions[1].height = 25

# Lưu file
output_path = r"d:\unity\Blade-Pursuit\Test_Execution_Report.xlsx"
wb.save(output_path)
print(f"Excel file created successfully: {output_path}")
print(f"Total test cases: {test_counter - 1}")
